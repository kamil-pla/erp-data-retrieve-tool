import os
import time
from webdriver_instance import driver
from dotenv import load_dotenv

from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import pandas as pd
import openpyxl

import win32clipboard
import pyautogui

tool_title = """\n
                   ---
                ---------
        --------------------------    
 ERP  data  retrieve  tool  from  JD Edwards
        --------------------------    
                ---------
                   ---

"""

# Pre-set branch codes:
branch1_code = 'branch1_code'
branch2_code = 'branch2_code'

# URL to the RMA requests output Excel file
url_output_file = "url_to_the_output_file"

# URL to JD Edwards
url_jd_edwards = "url_to_JD_Edwards"

# XPaths GENERAL
xpath_home ="/html/body/div[2]/div[5]/div[3]/div[4]/div[1]/div[2]/div/div[1]/table/tbody/tr/td[2]"
xpath_home_icon = "/html/body/div[2]/div[5]/div[3]/div[4]/div[1]/div[2]/div/div[1]/table/tbody/tr/td[1]/img"

# XPath for JDE login
xpath_jde_sign_in = "/html/body/div/table/tbody/tr[2]/td/form/table/tbody/tr/td/div/table/tbody/tr[2]/td[2]/table/" \
                    "tbody/tr[3]/td/div[14]/input"


# XPaths for Export Control

# XPath opening Export Control transaction
xpath_export_control = "/html/body/div[2]/div[5]/div[3]/div[4]/div[3]/div[2]/div/div[9]/table/tbody/tr/td[2]"

# XPath for part number field in the iFrame
xpath_iframe_cust_pn = '//*[@id="e1menuAppIframe"]'
xpath_branch1 = [
    "/html/body/form[3]/div/table/tbody/tr/td/div/span[10]/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[2]/td/div/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr/td[4]/div/input",
    "/html/body/form[3]/div/table/tbody/tr/td/div/span[10]/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[2]/td/div/table/tbody/tr/td[2]/table/tbody/tr[2]/td/table/tbody/tr/td[4]/div/input",
    "/html/body/form[3]/div/table/tbody/tr/td/div/span[10]/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[2]/td/div/table/tbody/tr/td[2]/table/tbody/tr[3]/td/table/tbody/tr/td[4]/div/input",
    "/html/body/form[3]/div/table/tbody/tr/td/div/span[10]/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[2]/td/div/table/tbody/tr/td[2]/table/tbody/tr[4]/td/table/tbody/tr/td[4]/div/input",
    "/html/body/form[3]/div/table/tbody/tr/td/div/span[10]/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[2]/td/div/table/tbody/tr/td[2]/table/tbody/tr[5]/td/table/tbody/tr/td[4]/div/input",
    "/html/body/form[3]/div/table/tbody/tr/td/div/span[10]/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[2]/td/div/table/tbody/tr/td[2]/table/tbody/tr[6]/td/table/tbody/tr/td[4]/div/input"
]
xpath_branch2 = [
    "/html/body/form[3]/div/table/tbody/tr/td/div/span[10]/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[2]/td/div/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr/td[5]/div/input",
    "/html/body/form[3]/div/table/tbody/tr/td/div/span[10]/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[2]/td/div/table/tbody/tr/td[2]/table/tbody/tr[2]/td/table/tbody/tr/td[5]/div/input",
    "/html/body/form[3]/div/table/tbody/tr/td/div/span[10]/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[2]/td/div/table/tbody/tr/td[2]/table/tbody/tr[3]/td/table/tbody/tr/td[5]/div/input",
    "/html/body/form[3]/div/table/tbody/tr/td/div/span[10]/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[2]/td/div/table/tbody/tr/td[2]/table/tbody/tr[4]/td/table/tbody/tr/td[5]/div/input",
    "/html/body/form[3]/div/table/tbody/tr/td/div/span[10]/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[2]/td/div/table/tbody/tr/td[2]/table/tbody/tr[5]/td/table/tbody/tr/td[5]/div/input",
    "/html/body/form[3]/div/table/tbody/tr/td/div/span[10]/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[2]/td/div/table/tbody/tr/td[2]/table/tbody/tr[6]/td/table/tbody/tr/td[5]/div/input"
]
xpath_cust_pn_empty = "/html/body/form[3]/div/table/tbody/tr/td/div/span[10]/table/tbody/tr[2]/td/table/tbody/" \
                      "tr/td/table/tbody/tr[2]/td/div/table/tbody/tr/td[2]/table/tbody/tr[6]/td/table/tbody/tr/td[3]/" \
                      "div/input"

xpath_clipboard = "/html/body/div/form/div/table/tbody/tr/td/table[1]/tbody/tr[5]/td/table/tbody/tr/td/table[2]/" \
                  "tbody/tr/td/table/tbody/tr[3]/td[1]/input"

xpath_apply = "/html/body/div[1]/form/div/table/tbody/tr/td/table[1]/tbody/tr[5]/td/table/tbody/tr/td/div[1]/input"

# XPaths for Unit Cost retrieve
xpath_po = "/html/body/div[2]/div[5]/div[3]/div[4]/div[3]/div[2]/div/div[4]/table/tbody/tr/td[1]/img"
xpath_po2 = "/html/body/div[2]/div[5]/div[3]/div[4]/div[3]/div[2]/div/div[4]/table/tbody/tr/td[2]"
xpath_iframe_po_input = '//*[@id="e1menuAppIframe"]'
xpath_po_input = '//*[@id="C0_13"]'
xpath_item_no_input = '//*[@id="C0_19"]'
xpath_iframe_bottom = "/html/body/div[2]/div[5]/div[2]/div[1]/div/iframe"
xpath_bottom_field = [
    "/html/body/form[3]/div/table/tbody/tr/td/div/span[2]/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[1]/td/"
    "div[1]/table/tbody/tr[1]/td[2]/div/nobr/input",
    "/html/body/form[3]/div/table/tbody/tr/td/div/span[2]/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[1]/td/"
    "div[1]/table/tbody/tr[1]/td[3]/div/nobr/input",
    "/html/body/form[3]/div/table/tbody/tr/td/div/span[2]/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[1]/td/"
    "div[1]/table/tbody/tr[1]/td[4]/div/nobr/input"
]
xpath_unit_price = "/html/body/form[3]/div/table/tbody/tr/td/div/span[2]/table/tbody/tr[2]/td/table/tbody/tr/td/" \
                   "table/tbody/tr[2]/td/div/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr/td[22]/div"
xpath_records_table = "/html/body/form[3]/div/table/tbody/tr/td/div/span[2]/table/tbody/tr[2]/td/table/tbody/tr/" \
                      "td/table/tbody/tr[2]/td/div/table/tbody/tr/td[2]/table/tbody/tr/td/table/tbody/tr/td[7]/div"

env_files = {
    '1': 'C:/Users/user1_name/.env', '2': 'C:/Users/user2_name/.env', '3': 'C:/Users/user3_name/.env',
    '4': 'C:/Users/user4_name/.env', '5': 'C:/Users/user5_name/.env', '6': 'C:/Users/user6_name/.env',
    '7': 'C:/Users/user7_name/.env', '8': 'C:/Users/user8_name/.env'
}


user_names = {
    '1': 'user1', '2': 'user2', '3': 'user3', '4': 'user4', '5': 'user5', '6': 'user6', '7': 'user7',
    '8': 'user8'
}


### GENERAL FUNCTIONS

def set_zoom(zoom_level):
    # setting zoom in order to load the table view correctly
    driver.get('chrome://settings/')
    driver.execute_script(f'chrome.settingsPrivate.setDefaultZoom({zoom_level});')


def get_username():
    user_welcome = ">>>>> \n\nPlease specify user to login into JDE or press 0 for EXIT: "
    prompt_username = "\nYour choice: "
    print(tool_title)
    print(user_welcome,"\n")
    for item in user_names.items():
        print(item[0],':', item[1])

    username_choice = input(prompt_username)

    if username_choice == '0':
        return '0'
    return username_choice


def get_password():
    global password_session
    password_session = pyautogui.password('Enter JD Edwards password: ')
    return password_session


username_choice = get_username()
password_session = get_password()


def set_dotenv():
    if username_choice in env_files:
        env_file_path = env_files.get(username_choice)
        print(f"Your credential file path is: {env_file_path}")
        return env_file_path
    elif username_choice == '0':
        exit()
    else:
        print(f"Cannot find .env file with credentials in the C:/Users/... folder")


curr_env_path = set_dotenv()
print(f"Searching for credentials in: {curr_env_path}")
load_dotenv(dotenv_path=curr_env_path)


def run_login_to_jde():
    login_to_jde(driver)


def click_xpath(driver, xpath):
    try:
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.presence_of_element_located((By.XPATH, xpath)))
        element.click()
    except TimeoutException:
        print("Timeout occurred while trying to open the link.")
        pass


def login_to_jde(driver):
    print(f"\n-----\nLogging into JDE as: ", os.getenv('user'))
    try:
        driver.get(url_jd_edwards)
    except TimeoutException:
        print("Timeout occurred while trying to open the link, opening in new window")
        driver.execute_script("window.open('url_jd_edwards')")
        driver.switch_to.window(driver.window_handles[-1])

    time.sleep(2)

    username = driver.find_element_by_id("User")
    password = driver.find_element_by_id("Password")
    username.send_keys(os.environ.get('user'))
    password.send_keys(password_session)
    click_xpath(driver, xpath_jde_sign_in)  # SignIn button
    time.sleep(2)


def define_range_of_rows():
    global request_row_no1, request_row_no2
    prompt1 = "\nEnter the first 'RMA requests' row number for searching: "
    prompt2 = "\nEnter the last 'RMA requests' row number for searching: "
    request_row_no1 = input(prompt1)
    request_row_no2 = input(prompt2)


def return_to_homepage():
    time.sleep(1)
    driver.switch_to.default_content()
    click_xpath(driver, xpath_home_icon)


def branch_code_change():
    global branch1_code
    branch1_code = input("Provide branch 1 code please: ")
    global branch2_code
    branch2_code = input("Provide branch 2 code please: ")
    return branch1_code, branch2_code


### EXPORT CONTROL FUNCTIONS


def run_export_control_script():
    print(f"\nRunning 'Export control' script. It should take less than a minute.")
    try:
        login_to_jde(driver)
        open_export_control(driver)
        retrieve_export_control()
        save_export_control()
        return_to_homepage()
        print("\nAll done. Process completed")
    except NoSuchElementException:
        open_export_control(driver)
        retrieve_export_control()
        save_export_control()
        return_to_homepage()
        print("\nAll done. Process completed")


def open_export_control(driver):
    while True:
        try:
            click_xpath(driver, xpath_export_control)
            break
        except:
            print("Apologies, I cannot open export control module")


def get_pn_data():
    path = url_output_file
    sheet_name = "Values_RMA_Requests"
    global row_number1
    row_number1 = int(request_row_no1) + 1
    global row_number2
    row_number2 = int(request_row_no2) + 1

    # Read the specified range of rows from the Excel file using Pandas
    df = pd.read_excel(path, sheet_name=sheet_name, header=None, skiprows=range(row_number1 - 1),
                       usecols="B", engine="openpyxl")
    # Extract the values from the specified range of rows
    pn_data = df.iloc[:row_number2 - row_number1 + 1, 0].tolist()
    print(f"\nSearching Excel file part numbers...")
    print(f"\nPart No. have been found:\n {pn_data}")
    return pn_data


def retrieve_export_control():
    pn_data = get_pn_data()
    # Calculate the number of lines based on the difference between row_number1 and row_number2
    number_of_lines = row_number2 - row_number1 + 1
    par = number_of_lines + 1
    # Preparing list of part number field XPaths which starting from first position of part no. xpath and then rest of
    # XPaths will be added based on parametres which amount = number of lines
    list_of_xpath_pn = ["/html/body/form[3]/div/table/tbody/tr/td/div/span[10]/table/tbody/tr[2]/td/table/tbody/tr/td/"
                      "table/tbody/tr[2]/td/div/table/tbody/tr/td[2]/table/tbody/tr/td/table/tbody/tr/td[3]/div/input"]

    for p in range(2, number_of_lines + 1):
        xpath_cust_pn = f"/html/body/form[3]/div/table/tbody/tr/td/div/span[10]/table/tbody/tr[2]/td/table/tbody/tr/" \
                        f"td/table/tbody/tr[2]/td/div/table/tbody/tr/td[2]/table/tbody/tr[{p}]/td/table/tbody/tr/" \
                        f"td[3]/div/input"
        list_of_xpath_pn.append(xpath_cust_pn)

    time.sleep(1)

    # Switch to the iframe containing the table element
    iframe_element = driver.find_element_by_xpath(xpath_iframe_cust_pn)
    driver.switch_to.frame(iframe_element)

    time.sleep(1)

    for pn, xpath_custpn, br1, br2 in zip(pn_data, list_of_xpath_pn, xpath_branch1, xpath_branch2):
        customer_pn = driver.find_element_by_xpath(xpath_custpn)
        customer_pn.send_keys(pn)

        branch1 = driver.find_element_by_xpath(br1)
        branch1.send_keys(branch1_code)

        branch2 = driver.find_element_by_xpath(br2)
        branch2.send_keys(branch2_code)
        branch2.send_keys(Keys.ENTER)

        time.sleep(.5)

    time.sleep(1)

    # xpath for an empty field below
    xpath_cust_pn_param = f"/html/body/form[3]/div/table/tbody/tr/td/div/span[10]/table/tbody/tr[2]/td/table/tbody/" \
                          f"tr/td/table/tbody/tr[2]/td/div/table/tbody/tr/td[2]/table/tbody/tr[{par}]/td/table/tbody/" \
                          f"tr/td[3]/div/input"
    custpn_empty = driver.find_element_by_xpath(xpath_cust_pn_param)

    # JDE hotkey for Export Grid Data
    custpn_empty.send_keys(Keys.CONTROL + Keys.SHIFT + 'E')

    time.sleep(2)

    data_to_clipboard = driver.find_element_by_xpath(xpath_clipboard)
    data_to_clipboard.click()

    apply_button = driver.find_element_by_xpath(xpath_apply)
    apply_button.click()
    time.sleep(.5)
    apply_button.send_keys(Keys.CONTROL + 'C')

    print(f"\nExport control data retrieved from JDE")


def save_export_control():
    # Get data from clipboard
    win32clipboard.OpenClipboard()
    clipboard_data = win32clipboard.GetClipboardData()
    win32clipboard.CloseClipboard()

    # Open Excel workbook
    url = url_output_file
    workbook = openpyxl.load_workbook(url)
    worksheet = workbook["Raw_JDE"]

    data = clipboard_data
    rows = data.split('\n')
    columns = [row.split('\t') for row in rows]

    # Find the first empty row in the "Raw_JDE" sheet
    empty_row = 1
    while worksheet.cell(row=empty_row, column=1).value is not None:
        empty_row += 1

    # Iterate through all rows starting from the second row
    for i in range(1, len(columns)):
        # Move to the next empty row
        empty_row += 1

    # Paste the row data into the empty row
        for j, column in enumerate(columns[i], start=1):
            worksheet.cell(row=empty_row, column=j, value=str(column).strip())

    time.sleep(0.5)

    # Save changes
    workbook.save(url)
    print(f"\nSaving data in the Excel file...")


### UNIT COST FUNCTIONS

def run_unit_cost_script():
    print(f"\nRunning 'Unit cost' script. It should take around "
          f"{(round((int(request_row_no2) - int(request_row_no1)),0)+1)*15} seconds.")
    try:
        set_zoom(.4)
        time.sleep(0.5)
        login_to_jde(driver)
        write_unit_costs()
        print("\nAll done. Process completed")
        set_zoom(.9)
        driver.back()
        return_to_homepage()
    except NoSuchElementException:
        open_po_enquiry(driver)
        write_unit_costs()
        print("\nAll done. Process completed")
        set_zoom(.9)
        driver.back()


def open_po_enquiry(driver):
    while True:
        try:
            driver.maximize_window()
            # PO enquiry enter
            click_xpath(driver, xpath_po)
            time.sleep(2)
            break
        except:
            print("Apologies, I cannot open PO enquiry search module. Let's try again...")
            click_xpath(driver, xpath_home_icon)
            time.sleep(2)
            click_xpath(driver, xpath_po2)
            time.sleep(1)


def switch_to_iframe(xpath1, xpath2):
    # switching to the iframes
    driver.switch_to.default_content()
    iframe_element = driver.find_element_by_xpath(xpath1)
    driver.switch_to.frame(iframe_element)

    move_to_iframe = driver.find_element_by_xpath(xpath2)
    move_to_iframe.click()


def get_po_order():
    print(f"\nSearching Excel file for Purchase Orders...")
    path = url_output_file
    sheet_name = "Values_RMA_Requests"
    row1 = int(row_number1)
    row2 = int(row_number2)

    # Read the Excel file from the bytes object using Pandas
    df = pd.read_excel(path, sheet_name=sheet_name, header=None, skiprows=range(row1 - 1), nrows=row2 - row1 + 1,
                       usecols="C", engine="openpyxl")
    po_data = df.iloc[:, 0].tolist()
    print("\nPurchase Orders found:")
    print(po_data)
    return po_data


def get_pn_and_po_combined():
    global pn_po_combined
    pn_po_combined = list(zip(get_pn_data(), get_po_order()))
    print (f"\nI have found following combinations of Part Numbers and Purchase Orders:\n {pn_po_combined}")
    return pn_po_combined


def po_enquiry_search(pn, po):
    # this open function has been put here because under run_unit_cost_script()
    # it did not erase previous PN input
    open_po_enquiry(driver)
    # PO input field
    switch_to_iframe(xpath_iframe_po_input, xpath_po_input)
    time.sleep(3)   # time delay adjusted after many tests

    po_input = driver.find_element_by_xpath(xpath_po_input)
    po_input.send_keys(Keys.BACKSPACE)
    po_input.send_keys(po)

    item_input = driver.find_element_by_xpath(xpath_item_no_input)
    item_input.send_keys(Keys.CONTROL + 'a')
    time.sleep(.5)
    item_input.send_keys(Keys.BACKSPACE)
    item_input.send_keys(pn)

    item_input.send_keys(Keys.CONTROL + Keys.ALT + 'i')
    time.sleep(1)
    driver.switch_to.default_content()
    time.sleep(2)


def unit_price_retrieve():
    time.sleep(1)
    # switch to the iframe containing the unit price element
    iframe_element = driver.find_element_by_xpath(xpath_iframe_bottom)
    driver.switch_to.frame(iframe_element)
    # click into first empty bottom field
    bottom_field = driver.find_element_by_xpath(xpath_bottom_field[0])
    bottom_field.click()
    time.sleep(2)

    unit_price_field = None
    while unit_price_field is None:
        # try to find the element with unit price
        try:
            unit_price_field = WebDriverWait(driver, .3).until(EC.visibility_of_element_located((By.XPATH,
                                                                                                 xpath_unit_price)))
        except TimeoutException:
            print (f"Waiting too long. JDE servers might be overloaded or system error - please check if PN and PO are "
                   f"correct and try later")
            time.sleep(5)
            break
        except:
            # if the element is not found, scroll the table to the right
            records_table = driver.find_element_by_xpath(xpath_records_table)
            records_table.send_keys(Keys.ARROW_RIGHT)
            time.sleep(.3)  # wait for the table to refresh

    # get the value of the table element
    try:
        unit_cost = unit_price_field.text
        driver.switch_to.default_content()
        return unit_cost
    except AttributeError:
        print("One of unit price could not be find in JDE / Purchase Order Enquiry  ")
        pass

def unit_cost_list_created():
    pn_po_list = get_pn_and_po_combined()
    unit_cost_list = []
    number_of_lines = 1

    for pn, po in pn_po_list:
        print(f"\nSearching for unit cost based on PO and part number...")
        po_enquiry_search(pn, po)
        unit_cost = unit_price_retrieve()
        unit_cost_list.append(unit_cost)
    print(f"\nUnit cost found: {unit_cost_list}")
    return unit_cost_list


def write_unit_costs():
    # create a new workbook object
    url = url_output_file
    workbook = openpyxl.load_workbook(url)
    worksheet = workbook["Raw_JDE2"]

    # get the list of values to write
    values_cost = unit_cost_list_created()
    values_pn_po = pn_po_combined

    # Find the first empty row in the "Raw_JDE" sheet
    empty_row = 1
    while worksheet.cell(row=empty_row, column=1).value is not None:
        empty_row += 1

    # Iterate over the tuple with PN and PO and assign each cell with data type - PN as text and PO as number
    for row, (text, num) in enumerate(values_pn_po, start=empty_row):
        worksheet.cell(row=row, column=1, value=text)
        worksheet.cell(row=row, column=2, value=num)

    for row, value in enumerate(values_cost[0:], start=empty_row):
        worksheet.cell(row=row, column=3, value=value)

    # Save the workbook to a file
    workbook.save(url)
